import streamlit as st
import mysql.connector
from mysql.connector import Error
import cv2
import numpy as np
import face_recognition
import base64
import json
from datetime import datetime, timedelta, date
import pandas as pd
from PIL import Image
import io
import os
from streamlit_webrtc import webrtc_streamer, VideoTransformerBase, RTCConfiguration
import av
import logging
import time
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell
from io import BytesIO
from dotenv import load_dotenv
import re

# Page configuration
st.set_page_config(
    page_title="Biometric Access System",
    page_icon="👤",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Camera locations configuration
CAMERA_LOCATIONS = {
    "main_entry": {
        "name": "Ana Giris",
        "access_type": "entry",
        "description": "Bina ana giriş kapısı"
    },
    "main_exit": {
        "name": "Ana Cikis", 
        "access_type": "exit",
        "description": "Bina ana çıkış kapısı"
    },
    "internal_door_1": {
        "name": "Ara Kapi 1",
        "access_type": "internal",
        "description": "İç bölüm ara kapısı 1"
    },
    "internal_door_2": {
        "name": "Ara Kapi 2", 
        "access_type": "internal",
        "description": "İç bölüm ara kapısı 2"
    },
}

# Database configuration from environment variables
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'database': os.getenv('DB_NAME', 'biometric_system'),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASSWORD', ''),
    'charset': os.getenv('DB_CHARSET', 'utf8mb4'),
    'collation': os.getenv('DB_COLLATION', 'utf8mb4_unicode_ci')
}

class DatabaseManager:
    """Database operations manager"""
    
    def __init__(self):
        self.connection = None
    
    def connect(self):
        """Establish database connection"""
        try:
            self.connection = mysql.connector.connect(**DB_CONFIG)
            return True
        except Error as e:
            st.error(f"Database connection failed: {e}")
            return False
    
    def disconnect(self):
        """Close database connection"""
        if self.connection and self.connection.is_connected():
            self.connection.close()
    
    def execute_query(self, query, params=None, fetch=False):
        """Execute database query"""
        try:
            if not self.connection or not self.connection.is_connected():
                if not self.connect():
                    return None
            
            cursor = self.connection.cursor(dictionary=True)
            cursor.execute(query, params or ())
            
            if fetch:
                result = cursor.fetchall()
            else:
                self.connection.commit()
                result = cursor.lastrowid
            
            cursor.close()
            return result
            
        except Error as e:
            st.error(f"Database query failed: {e}")
            return None
    
    def add_person(self, name, employee_id, department, phone, email, face_image_base64, face_embeddings):
        """Add a new person to the database"""
        try:
            query = """
            INSERT INTO persons (name, employee_id, department, phone, email, face_image, face_embeddings, embedding_count)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """
            
            embedding_count = len(face_embeddings) if face_embeddings else 0
            embeddings_json = json.dumps(face_embeddings) if face_embeddings else None
            
            params = (name, employee_id, department, phone, email, face_image_base64, embeddings_json, embedding_count)
            
            person_id = self.execute_query(query, params)
            return person_id
            
        except Exception as e:
            st.error(f"Failed to add person: {e}")
            return None
    
    def get_persons(self, active_only=True):
        """Get all persons from database"""
        query = "SELECT * FROM persons"
        if active_only:
            query += " WHERE is_active = TRUE"
        query += " ORDER BY created_at DESC"
        
        return self.execute_query(query, fetch=True)
    
    def get_person_by_id(self, person_id):
        """Get person by ID"""
        query = "SELECT * FROM persons WHERE id = %s"
        result = self.execute_query(query, (person_id,), fetch=True)
        return result[0] if result else None
    
    def update_person(self, person_id, **kwargs):
        """Update person information"""
        set_clauses = []
        params = []
        
        for key, value in kwargs.items():
            if value is not None:
                set_clauses.append(f"{key} = %s")
                params.append(value)
        
        if not set_clauses:
            return False
        
        query = f"UPDATE persons SET {', '.join(set_clauses)} WHERE id = %s"
        params.append(person_id)
        
        result = self.execute_query(query, params)
        return result is not None
    
    def deactivate_person(self, person_id):
        """Deactivate a person"""
        query = "UPDATE persons SET is_active = FALSE WHERE id = %s"
        result = self.execute_query(query, (person_id,))
        return result is not None
    
    def add_access_log(self, person_id, access_type, confidence_score, location="Ana Giriş", device_info=None, notes=None):
        """Add access log entry"""
        query = """
        INSERT INTO access_logs (person_id, access_type, confidence_score, location, device_info, notes)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        
        params = (person_id, access_type, confidence_score, location, device_info, notes)
        return self.execute_query(query, params)
    
    def get_access_logs(self, person_id=None, limit=100):
        """Get access logs"""
        query = """
        SELECT al.*, p.name, p.employee_id 
        FROM access_logs al 
        JOIN persons p ON al.person_id = p.id
        """
        params = []
        
        if person_id:
            query += " WHERE al.person_id = %s"
            params.append(person_id)
        
        query += " ORDER BY al.access_time DESC LIMIT %s"
        params.append(limit)
        
        return self.execute_query(query, params, fetch=True)
    
    def get_employee_attendance_data(self, person_id, start_date, end_date):
        """Get detailed attendance data for an employee in date range"""
        query = """
        SELECT al.*, p.name, p.employee_id 
        FROM access_logs al 
        JOIN persons p ON al.person_id = p.id
        WHERE al.person_id = %s 
        AND DATE(al.access_time) BETWEEN %s AND %s
        ORDER BY al.access_time ASC
        """
        
        return self.execute_query(query, (person_id, start_date, end_date), fetch=True)
    
    def get_all_employees_attendance(self, start_date, end_date):
        """Get attendance data for all employees in date range"""
        query = """
        SELECT al.*, p.name, p.employee_id, p.department
        FROM access_logs al 
        JOIN persons p ON al.person_id = p.id
        WHERE DATE(al.access_time) BETWEEN %s AND %s
        ORDER BY p.name, al.access_time ASC
        """
        
        return self.execute_query(query, (start_date, end_date), fetch=True)
    
    def delete_person(self, person_id):
        """Delete a person and all related logs from the database"""
        try:
            # Önce access_logs'tan sil
            self.execute_query("DELETE FROM access_logs WHERE person_id = %s", (person_id,))
            # Sonra persons'tan sil
            self.execute_query("DELETE FROM persons WHERE id = %s", (person_id,))
            return True
        except Exception as e:
            st.error(f"Failed to delete person: {e}")
            return False

class FaceProcessor:
    """Face detection and recognition processor"""
    
    @staticmethod
    def detect_faces(image):
        """Detect faces in image and return face locations"""
        try:
            # Convert PIL image to numpy array if needed
            if isinstance(image, Image.Image):
                image = np.array(image)
            
            # Convert BGR to RGB if needed
            if len(image.shape) == 3 and image.shape[2] == 3:
                image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            else:
                image_rgb = image
            
            # Detect face locations
            face_locations = face_recognition.face_locations(image_rgb)
            return face_locations
            
        except Exception as e:
            logger.error(f"Face detection failed: {e}")
            return []
    
    @staticmethod
    def extract_embeddings(image, face_locations=None):
        """Extract face embeddings from image"""
        try:
            # Convert PIL image to numpy array if needed
            if isinstance(image, Image.Image):
                image = np.array(image)
            
            # Convert BGR to RGB if needed
            if len(image.shape) == 3 and image.shape[2] == 3:
                image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            else:
                image_rgb = image
            
            # Get face locations if not provided
            if face_locations is None:
                face_locations = face_recognition.face_locations(image_rgb)
            
            # Extract encodings
            face_encodings = face_recognition.face_encodings(image_rgb, face_locations)
            
            # Convert to list for JSON serialization
            embeddings = [encoding.tolist() for encoding in face_encodings]
            
            return embeddings, face_locations
            
        except Exception as e:
            logger.error(f"Embedding extraction failed: {e}")
            return [], []
    
    @staticmethod
    def image_to_base64(image):
        """Convert image to base64 string"""
        try:
            if isinstance(image, np.ndarray):
                # Convert numpy array to PIL Image
                if len(image.shape) == 3:
                    image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
                image = Image.fromarray(image)
            
            # Convert to base64
            buffered = io.BytesIO()
            image.save(buffered, format="JPEG", quality=85)
            img_str = base64.b64encode(buffered.getvalue()).decode()
            return img_str
            
        except Exception as e:
            logger.error(f"Image to base64 conversion failed: {e}")
            return None
    
    @staticmethod
    def base64_to_image(base64_str):
        """Convert base64 string to PIL Image"""
        try:
            img_data = base64.b64decode(base64_str)
            image = Image.open(io.BytesIO(img_data))
            return image
        except Exception as e:
            logger.error(f"Base64 to image conversion failed: {e}")
            return None

class CameraTransformer(VideoTransformerBase):
    """Kamera çekimi için özel video transformer"""
    
    def __init__(self):
        self.captured_frame = None
        self.should_capture = False
        self.frame_count = 0
    
    def recv(self, frame):
        img = frame.to_ndarray(format="bgr24")
        
        # Frame sayacını artır
        self.frame_count += 1
        
        # Eğer capture flag set ise, frame'i kaydet
        if self.should_capture:
            self.captured_frame = img.copy()
            self.should_capture = False
            
        # Yüz tespiti yap ve göster
        face_locations = FaceProcessor.detect_faces(img)
        
        # Tespit edilen yüzleri çerçevele
        for (top, right, bottom, left) in face_locations:
            cv2.rectangle(img, (left, top), (right, bottom), (0, 255, 0), 2)
            cv2.putText(img, f"Face Detected", (left, top - 10),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
        
        # Capture butonu için bilgi göster
        if face_locations:
            cv2.putText(img, "Face ready for capture", (10, 30),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
        else:
            cv2.putText(img, "No face detected", (10, 30),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
        
        return av.VideoFrame.from_ndarray(img, format="bgr24")
    
    def capture_frame(self):
        """Frame capture'ı tetikle"""
        self.should_capture = True
    
    def get_captured_frame(self):
        """Capture edilen frame'i döndür"""
        return self.captured_frame

class AttendanceAnalyzer:
    """Attendance and working hours analysis"""
    
    WORK_START_TIME = "08:00"
    WORK_END_TIME = "17:00"
    WORK_DAYS = [0, 1, 2, 3, 4]  # Monday to Friday (0=Monday, 6=Sunday)
    
    @staticmethod
    def is_workday(date_obj):
        """Check if given date is a workday"""
        return date_obj.weekday() in AttendanceAnalyzer.WORK_DAYS
    
    @staticmethod
    def get_workdays_in_range(start_date, end_date):
        """Get all workdays in date range"""
        workdays = []
        current_date = start_date
        while current_date <= end_date:
            if AttendanceAnalyzer.is_workday(current_date):
                workdays.append(current_date)
            current_date += timedelta(days=1)
        return workdays
    
    @staticmethod
    def analyze_employee_attendance(attendance_data, start_date, end_date):
        """Analyze attendance data for an employee"""
        if not attendance_data:
            return None
        
        # Group by date
        daily_data = {}
        for record in attendance_data:
            date_key = record['access_time'].date()
            if date_key not in daily_data:
                daily_data[date_key] = {'entries': [], 'exits': []}
            
            if record['access_type'] == 'entry':
                daily_data[date_key]['entries'].append(record['access_time'])
            elif record['access_type'] == 'exit':
                daily_data[date_key]['exits'].append(record['access_time'])
        
        # Analyze each workday
        analysis_results = []
        workdays = AttendanceAnalyzer.get_workdays_in_range(start_date, end_date)
        
        for workday in workdays:
            day_data = daily_data.get(workday, {'entries': [], 'exits': []})
            
            # Get earliest entry and latest exit
            first_entry = min(day_data['entries']) if day_data['entries'] else None
            last_exit = max(day_data['exits']) if day_data['exits'] else None
            
            # Calculate delays and early departures
            entry_delay = None
            exit_early = None
            working_hours = None
            status = "Absent"
            
            if first_entry and last_exit:
                # Calculate entry delay
                expected_entry = datetime.combine(workday, datetime.strptime(AttendanceAnalyzer.WORK_START_TIME, "%H:%M").time())
                if first_entry > expected_entry:
                    entry_delay = (first_entry - expected_entry).total_seconds() / 60  # minutes
                
                # Calculate early exit
                expected_exit = datetime.combine(workday, datetime.strptime(AttendanceAnalyzer.WORK_END_TIME, "%H:%M").time())
                if last_exit < expected_exit:
                    exit_early = (expected_exit - last_exit).total_seconds() / 60  # minutes
                
                # Calculate working hours
                working_hours = (last_exit - first_entry).total_seconds() / 3600  # hours
                status = "Present"
            
            elif first_entry:
                # Only entry, no exit
                expected_entry = datetime.combine(workday, datetime.strptime(AttendanceAnalyzer.WORK_START_TIME, "%H:%M").time())
                if first_entry > expected_entry:
                    entry_delay = (first_entry - expected_entry).total_seconds() / 60
                status = "Partial (No Exit)"
            
            elif last_exit:
                # Only exit, no entry
                expected_exit = datetime.combine(workday, datetime.strptime(AttendanceAnalyzer.WORK_END_TIME, "%H:%M").time())
                if last_exit < expected_exit:
                    exit_early = (expected_exit - last_exit).total_seconds() / 60
                status = "Partial (No Entry)"
            
            analysis_results.append({
                'date': workday,
                'status': status,
                'first_entry': first_entry,
                'last_exit': last_exit,
                'entry_delay_minutes': entry_delay,
                'exit_early_minutes': exit_early,
                'working_hours': working_hours,
                'expected_entry': datetime.combine(workday, datetime.strptime(AttendanceAnalyzer.WORK_START_TIME, "%H:%M").time()),
                'expected_exit': datetime.combine(workday, datetime.strptime(AttendanceAnalyzer.WORK_END_TIME, "%H:%M").time())
            })
        
        return analysis_results
    
    @staticmethod
    def generate_monthly_summary(analysis_results):
        """Generate summary statistics for monthly report"""
        if not analysis_results:
            return None
        
        total_workdays = len(analysis_results)
        present_days = sum(1 for day in analysis_results if day['status'] == 'Present')
        absent_days = sum(1 for day in analysis_results if day['status'] == 'Absent')
        partial_days = sum(1 for day in analysis_results if 'Partial' in day['status'])
        
        # Calculate average delays and early exits
        entry_delays = [day['entry_delay_minutes'] for day in analysis_results if day['entry_delay_minutes']]
        exit_earlies = [day['exit_early_minutes'] for day in analysis_results if day['exit_early_minutes']]
        
        avg_entry_delay = sum(entry_delays) / len(entry_delays) if entry_delays else 0
        avg_exit_early = sum(exit_earlies) / len(exit_earlies) if exit_earlies else 0
        
        # Calculate total delays and early exits
        total_entry_delay = sum(entry_delays) if entry_delays else 0
        total_exit_early = sum(exit_earlies) if exit_earlies else 0
        
        return {
            'total_workdays': total_workdays,
            'present_days': present_days,
            'absent_days': absent_days,
            'partial_days': partial_days,
            'attendance_rate': (present_days / total_workdays * 100) if total_workdays > 0 else 0,
            'avg_entry_delay_minutes': avg_entry_delay,
            'avg_exit_early_minutes': avg_exit_early,
            'total_entry_delay_minutes': total_entry_delay,
            'total_exit_early_minutes': total_exit_early,
            'total_working_hours': sum(day['working_hours'] for day in analysis_results if day['working_hours']) if analysis_results else 0
        }

class VideoTransformer(VideoTransformerBase):
    """Video transformer for recognition"""
    
    def __init__(self, camera_location="main_entry"):
        self.captured_frame = None
        self.last_recognition = None
        self.last_confidence = None
        self.db = None
        self.access_logged = False
        self.camera_location = camera_location
        self.last_person_id = None  # Track last recognized person to avoid duplicate logs
    
    def set_db(self, db):
        self.db = db
    
    def recv(self, frame):
        img = frame.to_ndarray(format="bgr24")
        face_locations = FaceProcessor.detect_faces(img)
        recognized_faces = []  # Store recognition results for each face
        
        if face_locations:
            embeddings, _ = FaceProcessor.extract_embeddings(img, face_locations)
            if embeddings and self.db is not None:
                persons = self.db.get_persons()
                
                # Process each detected face separately
                for face_idx, test_embedding in enumerate(embeddings):
                    best_match = None
                    best_confidence = 0
                    
                    # Compare this face with all persons in database
                    for person in persons:
                        if not person['face_embeddings']:
                            continue
                        try:
                            stored_embeddings = json.loads(person['face_embeddings'])
                            min_distance = float('inf')
                            
                            # Find best match for this person's embeddings
                            for stored_embedding in stored_embeddings:
                                distance = face_recognition.face_distance(
                                    [np.array(stored_embedding)], np.array(test_embedding)
                                )[0]
                                min_distance = min(min_distance, distance)
                            
                            # Check if this person is a better match
                            if min_distance < 0.6:
                                confidence = 1 - min_distance
                                if confidence > best_confidence:
                                    best_confidence = confidence
                                    best_match = person
                                    
                        except Exception as e:
                            logger.error(f"Error processing person {person['id']}: {e}")
                    
                    # Store result for this face
                    if best_match:
                        recognized_faces.append({
                            'face_idx': face_idx,
                            'person': best_match,
                            'confidence': best_confidence
                        })
                        
                        # Access log ekle (her frame için tekrar tekrar eklememek için person_id kontrolü)
                        current_person_id = best_match['id']
                        location_info = CAMERA_LOCATIONS.get(self.camera_location, CAMERA_LOCATIONS["main_entry"])
                        access_type = location_info["access_type"]
                        # Internal kapılar için log ekleme
                        if access_type == "internal":
                            continue
                        # Threshold ile log ekle (10 saniye)
                        now = time.time()
                        if not hasattr(self, 'last_log_times'):
                            self.last_log_times = {}
                        key = f"{current_person_id}_{self.camera_location}"
                        last_time = self.last_log_times.get(key, 0)
                        if now - last_time >= 10:
                            self.db.add_access_log(
                                person_id=current_person_id,
                                access_type=access_type,
                                confidence_score=best_confidence,
                                location=location_info["name"],
                                device_info=f"Camera: {location_info['description']}"
                            )
                            self.last_log_times[key] = now
                        self.access_logged = True
                        self.last_person_id = current_person_id
                
                # Update last recognition for display purposes (use the first recognized face)
                if recognized_faces:
                    self.last_recognition = recognized_faces[0]['person']
                    self.last_confidence = recognized_faces[0]['confidence']
                else:
                    self.last_recognition = None
                    self.last_confidence = None
                    self.access_logged = False
                    self.last_person_id = None
        
        # Yüzleri çerçevele ve her birine doğru label ekle
        for i, (top, right, bottom, left) in enumerate(face_locations):
            # Find if this face was recognized
            face_recognition_result = None
            for rec_face in recognized_faces:
                if rec_face['face_idx'] == i:
                    face_recognition_result = rec_face
                    break
            
            # Set label based on recognition result
            if face_recognition_result:
                person = face_recognition_result['person']
                confidence = face_recognition_result['confidence']
                label = f"{person['name']} ({confidence:.1%})"
                color = (0, 255, 0)  # Green for recognized
            else:
                label = "Unknown Face"
                color = (0, 0, 255)  # Red for unknown
            
            cv2.rectangle(img, (left, top), (right, bottom), color, 2)
            cv2.putText(img, label, (left, top - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)
        
        # Show access granted message if any face was recognized
        if recognized_faces:
            location_info = CAMERA_LOCATIONS.get(self.camera_location, CAMERA_LOCATIONS["main_entry"])
            access_type = location_info["access_type"]
            
            if access_type == "entry":
                message = "Giris Izni Verildi"
                color = (0, 255, 0)  # Green
            elif access_type == "exit":
                message = "Cikis Izni Verildi"
                color = (255, 165, 0)  # Orange
            else:  # internal
                message = "Ic Kapi Acildi"
                color = (255, 255, 0)  # Yellow
            
            cv2.putText(img, message, (10, 60),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, color, 3)
            
            # Show camera location
            cv2.putText(img, location_info["name"], (10, 100),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
        
        return av.VideoFrame.from_ndarray(img, format="bgr24")

# Initialize database manager
@st.cache_resource
def get_db_manager():
    return DatabaseManager()

# Yardımcı fonksiyon: dakika cinsinden süreyi saat, dakika, saniye olarak formatla
def format_minutes_to_hms(minutes):
    if minutes is None:
        return "0 sn"
    total_seconds = int(round(minutes * 60))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    parts = []
    if hours > 0:
        parts.append(f"{hours} saat")
    if minutes > 0:
        parts.append(f"{minutes} dakika")
    if seconds > 0 or not parts:
        parts.append(f"{seconds} saniye")
    return " ".join(parts)

def main():
    st.title("🔐 Biometric Access System")
    st.markdown("---")
    
    # Initialize database
    db = get_db_manager()
    
    # Sidebar navigation
    st.sidebar.title("Navigation")
    pages = [
        "👥 Add Person",
        "📋 View Persons",
        "📊 Access Logs",
        "📈 Attendance Reports",
        "🔍 Face Recognition Test"
    ]
    if 'selected_page' not in st.session_state:
        st.session_state.selected_page = pages[0]
    for p in pages:
        if st.sidebar.button(p, key=p):
            st.session_state.selected_page = p
    page = st.session_state.selected_page
    
    if page == "👥 Add Person":
        add_person_page(db)
    elif page == "📋 View Persons":
        view_persons_page(db)
    elif page == "📊 Access Logs":
        access_logs_page(db)
    elif page == "📈 Attendance Reports":
        attendance_reports_page(db)
    elif page == "🔍 Face Recognition Test":
        face_recognition_test_page(db)

def add_person_page(db):
    """Page for adding new persons"""
    st.header("👥 Add New Person")
    
    # Initialize session state for captured image
    if 'captured_image' not in st.session_state:
        st.session_state.captured_image = None
        st.session_state.face_embeddings = []
        st.session_state.face_image_base64 = None
    
    # Form for person details
    with st.form("add_person_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            name = st.text_input("Full Name *", placeholder="Enter full name")
            employee_id = st.text_input("Employee ID *", placeholder="Enter employee ID")
            department = st.text_input("Department", placeholder="Enter department")
        
        with col2:
            phone = st.text_input("Phone Number", placeholder="Enter phone number")
            email = st.text_input("Email", placeholder="Enter email address")
        
        st.markdown("### 📸 Photo Upload")
        
        # Photo upload options
        photo_option = st.radio(
            "Choose photo source:",
            ["📁 Upload from file", "📷 Capture from camera"]
        )
        
        uploaded_image = None
        face_image_base64 = None
        face_embeddings = []
        
        if photo_option == "📁 Upload from file":
            uploaded_file = st.file_uploader(
                "Choose an image file",
                type=['jpg', 'jpeg', 'png'],
                help="Upload a clear photo of the person's face"
            )
            
            if uploaded_file is not None:
                uploaded_image = Image.open(uploaded_file)
                st.image(uploaded_image, caption="Uploaded Image", width=300)
                
                # Process the image
                with st.spinner("Processing image..."):
                    face_locations = FaceProcessor.detect_faces(uploaded_image)
                    
                    if face_locations:
                        st.success(f"✅ Detected {len(face_locations)} face(s)")
                        
                        # Extract embeddings
                        embeddings, _ = FaceProcessor.extract_embeddings(uploaded_image, face_locations)
                        face_embeddings = embeddings
                        
                        # Convert to base64
                        face_image_base64 = FaceProcessor.image_to_base64(uploaded_image)
                        
                        # Show detected faces
                        img_array = np.array(uploaded_image)
                        for i, (top, right, bottom, left) in enumerate(face_locations):
                            cv2.rectangle(img_array, (left, top), (right, bottom), (0, 255, 0), 2)
                            cv2.putText(img_array, f"Face {i+1}", (left, top - 10), 
                                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)
                        
                        st.image(img_array, caption="Detected Faces", width=300)
                    else:
                        st.error("❌ No faces detected in the image. Please upload a clearer photo.")
        
        # Submit button
        submitted = st.form_submit_button("➕ Add Person", type="primary")
    
    # Camera capture functionality (outside the form)
    if photo_option == "📷 Capture from camera":
        st.info("📷 Kamerayı açmak ve fotoğraf çekmek için aşağıdaki butonları kullanın.")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("📷 Open Camera", type="primary"):
                st.session_state.camera_open = True
        
        with col2:
            if st.button("📸 Take Photo"):
                if hasattr(st.session_state, 'camera_open') and st.session_state.camera_open:
                    # Simple camera capture using OpenCV
                    cap = cv2.VideoCapture(0)
                    if cap.isOpened():
                        # Set camera properties for better quality
                        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
                        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
                        cap.set(cv2.CAP_PROP_AUTOFOCUS, 1)
                        cap.set(cv2.CAP_PROP_AUTO_EXPOSURE, 0.25)  # Enable auto exposure
                        
                        # Wait for camera to initialize and adjust
                        st.info("📷 Kamera ayarlanıyor, lütfen bekleyin...")
                        time.sleep(2)  # Wait 2 seconds for camera to adjust
                        
                        # Read multiple frames to let camera adjust
                        for _ in range(5):
                            cap.read()
                            time.sleep(0.1)
                        
                        # Now take the actual photo
                        ret, frame = cap.read()
                        if ret:
                            # Convert BGR to RGB
                            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                            captured_image = Image.fromarray(frame_rgb)
                            
                            # Store in session state
                            st.session_state.captured_image = captured_image
                            
                            # Process the image
                            with st.spinner("Processing captured image..."):
                                face_locations = FaceProcessor.detect_faces(captured_image)
                                
                                if face_locations:
                                    st.success(f"✅ {len(face_locations)} yüz algılandı!")
                                    
                                    # Extract embeddings
                                    embeddings, _ = FaceProcessor.extract_embeddings(captured_image, face_locations)
                                    st.session_state.face_embeddings = embeddings
                                    
                                    # Convert to base64
                                    st.session_state.face_image_base64 = FaceProcessor.image_to_base64(captured_image)
                                    
                                    # Show detected faces
                                    img_with_faces = np.array(captured_image)
                                    for i, (top, right, bottom, left) in enumerate(face_locations):
                                        cv2.rectangle(img_with_faces, (left, top), (right, bottom), (0, 255, 0), 2)
                                        cv2.putText(img_with_faces, f"Face {i+1}", (left, top - 10), 
                                                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)
                                    
                                    st.image(img_with_faces, caption="Captured Photo with Detected Faces", width=300)
                                else:
                                    st.error("❌ Yüz algılanamadı. Tekrar deneyin.")
                        else:
                            st.error("❌ Fotoğraf çekilemedi.")
                        cap.release()
                    else:
                        st.error("❌ Kamera açılamadı.")
                else:
                    st.warning("⚠️ Önce kamerayı açın!")
        
        with col3:
            if st.button("🔄 Reset"):
                st.session_state.captured_image = None
                st.session_state.face_embeddings = []
                st.session_state.face_image_base64 = None
                st.session_state.camera_open = False
                st.rerun()
        
        # Show captured image if available
        if st.session_state.captured_image is not None:
            st.success("✅ Fotoğraf çekildi ve işlendi!")
            st.image(st.session_state.captured_image, caption="Captured Photo", width=300)
            
            if st.session_state.face_embeddings:
                st.info(f"📊 {len(st.session_state.face_embeddings)} yüz embedding'i hazır.")
    
    # Handle form submission
    if submitted:
        # Get the appropriate image data based on photo option
        if photo_option == "📁 Upload from file":
            final_face_embeddings = face_embeddings
            final_face_image_base64 = face_image_base64
        else:  # Camera capture
            final_face_embeddings = st.session_state.face_embeddings
            final_face_image_base64 = st.session_state.face_image_base64
        
        # Validation
        if not name or not employee_id:
            st.error("❌ Name and Employee ID are required!")
            return
        
        if not final_face_embeddings:
            st.error("❌ Please upload or capture a photo with a detectable face!")
            return
        
        # Add person to database
        with st.spinner("Adding person to database..."):
            person_id = db.add_person(
                name=name,
                employee_id=employee_id,
                department=department or None,
                phone=phone or None,
                email=email or None,
                face_image_base64=final_face_image_base64,
                face_embeddings=final_face_embeddings
            )
        
        if person_id:
            st.success(f"✅ Person added successfully! ID: {person_id}")
            st.balloons()
            
            # Clear session state
            st.session_state.captured_image = None
            st.session_state.face_embeddings = []
            st.session_state.face_image_base64 = None
            st.session_state.camera_open = False
            
            # Show summary
            st.markdown("### 📋 Summary")
            col1, col2 = st.columns(2)
            with col1:
                st.write(f"**Name:** {name}")
                st.write(f"**Employee ID:** {employee_id}")
                st.write(f"**Department:** {department or 'N/A'}")
            with col2:
                st.write(f"**Phone:** {phone or 'N/A'}")
                st.write(f"**Email:** {email or 'N/A'}")
                st.write(f"**Face Embeddings:** {len(final_face_embeddings)}")
        else:
            st.error("❌ Failed to add person. Please try again.")

def view_persons_page(db):
    """Page for viewing all persons"""
    st.header("📋 Registered Persons")
    
    # Get persons from database
    persons = db.get_persons()
    
    if not persons:
        st.info("No persons registered yet.")
        return
    
    # Display statistics
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Persons", len(persons))
    with col2:
        active_count = sum(1 for p in persons if p['is_active'])
        st.metric("Active", active_count)
    with col3:
        inactive_count = len(persons) - active_count
        st.metric("Inactive", inactive_count)
    with col4:
        total_embeddings = sum(p['embedding_count'] or 0 for p in persons)
        st.metric("Total Embeddings", total_embeddings)
    
    st.markdown("---")
    
    # Search and filter
    col1, col2 = st.columns(2)
    with col1:
        search_term = st.text_input("🔍 Search by name or employee ID")
    with col2:
        status_filter = st.selectbox("Filter by status", ["All", "Active", "Inactive"])
    
    # Filter persons
    filtered_persons = persons
    
    if search_term:
        filtered_persons = [
            p for p in filtered_persons 
            if search_term.lower() in p['name'].lower() or 
               search_term.lower() in p['employee_id'].lower()
        ]
    
    if status_filter != "All":
        is_active = status_filter == "Active"
        filtered_persons = [p for p in filtered_persons if p['is_active'] == is_active]
    
    # Display persons in cards
    for person in filtered_persons:
        with st.expander(f"👤 {person['name']} ({person['employee_id']})"):
            col1, col2, col3 = st.columns([1, 2, 1])
            
            with col1:
                # Display photo if available
                if person['face_image']:
                    try:
                        image = FaceProcessor.base64_to_image(person['face_image'])
                        if image:
                            st.image(image, width=150, caption="Photo")
                    except:
                        st.write("📷 Photo unavailable")
                else:
                    st.write("📷 No photo")
            
            with col2:
                st.write(f"**Employee ID:** {person['employee_id']}")
                st.write(f"**Department:** {person['department'] or 'N/A'}")
                st.write(f"**Phone:** {person['phone'] or 'N/A'}")
                st.write(f"**Email:** {person['email'] or 'N/A'}")
                st.write(f"**Face Embeddings:** {person['embedding_count'] or 0}")
                st.write(f"**Created:** {person['created_at']}")
                st.write(f"**Status:** {'🟢 Active' if person['is_active'] else '🔴 Inactive'}")
            
            with col3:
                if st.button(f"🔍 View Logs", key=f"logs_{person['id']}"):
                    st.session_state['view_logs_person_id'] = person['id']
                
                if person['is_active']:
                    if st.button(f"🚫 Deactivate", key=f"deactivate_{person['id']}"):
                        if db.deactivate_person(person['id']):
                            st.success("Person deactivated!")
                            st.rerun()
                        else:
                            st.error("Failed to deactivate person!")
                # Delete Person butonu ve onay
                if st.button(f"🗑️ Delete Person", key=f"delete_{person['id']}"):
                    st.session_state[f"confirm_delete_{person['id']}"] = True
                if st.session_state.get(f"confirm_delete_{person['id']}", False):
                    st.warning(f"Are you sure you want to delete {person['name']} ({person['employee_id']})? This cannot be undone.")
                    col_yes, col_no = st.columns(2)
                    with col_yes:
                        if st.button(f"Yes, delete {person['name']}", key=f"yes_delete_{person['id']}"):
                            if db.delete_person(person['id']):
                                st.success("Person and all related data deleted!")
                                st.session_state[f"confirm_delete_{person['id']}"] = False
                                st.rerun()
                            else:
                                st.error("Failed to delete person!")
                                st.session_state[f"confirm_delete_{person['id']}"] = False
                    with col_no:
                        if st.button("Cancel", key=f"cancel_delete_{person['id']}"):
                            st.session_state[f"confirm_delete_{person['id']}"] = False

def access_logs_page(db):
    """Page for viewing access logs"""
    st.header("📊 Access Logs")
    
    # Filters
    persons_list = db.get_persons()
    person_names = [p['name'] for p in persons_list]
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        with st.expander("Çalışan(lar) ile filtrele"):
            selected_persons = []
            for name in person_names:
                if st.checkbox(name, key=f"person_checkbox_{name}"):
                    selected_persons.append(name)
    
    with col2:
        # Updated access type filter with new types
        access_type_options = ["All", "entry", "exit", "internal"]
        access_type_filter = st.selectbox("Access type", access_type_options)
    
    with col3:
        # Add location filter
        location_options = ["All"] + [info["name"] for info in CAMERA_LOCATIONS.values()]
        location_filter = st.selectbox("Location", location_options)
    
    with col4:
        limit = st.number_input("Number of records", min_value=10, max_value=1000, value=100)
    
    # Get logs
    if not selected_persons:
        # Hiç kimse seçili değilse, tüm çalışanların loglarını göster
        selected_ids = [p['id'] for p in persons_list]
    else:
        selected_ids = [p['id'] for p in persons_list if p['name'] in selected_persons]
    
    logs = []
    for pid in selected_ids:
        logs += db.get_access_logs(pid, limit)
    
    # Tekrar eden kayıtları önle (aynı log birden fazla seçilirse)
    logs = {log['id']: log for log in logs}.values() if logs else []
    logs = list(logs)
    
    if not logs:
        st.info("No access logs found.")
        return
    
    # Filter by access type
    if access_type_filter != "All":
        logs = [log for log in logs if log['access_type'] == access_type_filter]
    
    # Filter by location
    if location_filter != "All":
        logs = [log for log in logs if log['location'] == location_filter]
    
    # Display statistics
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.metric("Total Records", len(logs))
    with col2:
        entry_count = sum(1 for log in logs if log['access_type'] == 'entry')
        st.metric("Entries", entry_count)
    with col3:
        exit_count = sum(1 for log in logs if log['access_type'] == 'exit')
        st.metric("Exits", exit_count)
    with col4:
        internal_count = sum(1 for log in logs if log['access_type'] == 'internal')
        st.metric("Internal", internal_count)
    with col5:
        avg_confidence = sum(log['confidence_score'] or 0 for log in logs) / len(logs) if logs else 0
        st.metric("Avg Confidence", f"{avg_confidence:.3f}")
    
    # Display logs table
    if logs:
        df = pd.DataFrame(logs)
        
        # Format the dataframe
        df['access_time'] = pd.to_datetime(df['access_time'])
        df['confidence_score'] = df['confidence_score'].fillna(0).round(3)
        
        # Select columns to display
        display_columns = ['access_time', 'name', 'employee_id', 'access_type', 
                         'confidence_score', 'location', 'device_info']
        df_display = df[display_columns].copy()
        
        # Rename columns for better display
        df_display.columns = ['Time', 'Name', 'Employee ID', 'Type', 
                            'Confidence', 'Location', 'Device']
        
        # Add color coding for access types
        def color_access_type(val):
            if val == 'entry':
                return 'background-color: #d4edda'  # Light green
            elif val == 'exit':
                return 'background-color: #fff3cd'  # Light yellow
            elif val == 'internal':
                return 'background-color: #d1ecf1'  # Light blue
            return ''
        
        # Apply styling
        styled_df = df_display.style.applymap(color_access_type, subset=['Type'])
        
        st.dataframe(styled_df, use_container_width=True)
        
        # Download options
        st.markdown("### 📥 Download Options")
        col1, col2 = st.columns(2)
        
        with col1:
            # Download as CSV (without colors)
            csv = df_display.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="📄 Download as CSV",
                data=csv,
                file_name=f"access_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
        
        with col2:
            # Download as Excel with colors
            try:
                # Create Excel workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Access Logs"
                
                # Add data to worksheet
                for r in dataframe_to_rows(df_display, index=False, header=True):
                    ws.append(r)
                
                # Apply colors to Type column (column D)
                entry_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # Light green
                exit_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")   # Light yellow
                internal_fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid") # Light blue
                
                # Apply colors to cells based on access type
                for row in range(2, len(df_display) + 2):  # Start from row 2 (skip header)
                    cell_value = ws[f'D{row}'].value
                    if cell_value == 'entry':
                        ws[f'D{row}'].fill = entry_fill
                    elif cell_value == 'exit':
                        ws[f'D{row}'].fill = exit_fill
                    elif cell_value == 'internal':
                        ws[f'D{row}'].fill = internal_fill
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = None
                    for cell in column:
                        if not isinstance(cell, MergedCell):
                            if column_letter is None:
                                column_letter = cell.column_letter
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                    if column_letter:
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to BytesIO
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                st.download_button(
                    label="📊 Download as Excel (with colors)",
                    data=excel_buffer.getvalue(),
                    file_name=f"access_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            except ImportError:
                st.warning("⚠️ Excel download requires openpyxl. Install with: `pip install openpyxl`")
                st.info("📄 CSV download is still available.")

def attendance_reports_page(db):
    """Page for attendance reports and analysis"""
    st.header("📈 Attendance Reports")
    
    # Get all employees
    employees = db.get_persons()
    if not employees:
        st.info("No employees found. Please add employees first.")
        return
    
    # Report type selection
    report_type = st.selectbox(
        "Select Report Type",
        ["Individual Employee Report", "Monthly Summary Report", "Department Report"]
    )
    
    if report_type == "Individual Employee Report":
        individual_employee_report(db, employees)
    elif report_type == "Monthly Summary Report":
        monthly_summary_report(db, employees)
    elif report_type == "Department Report":
        department_report(db, employees)

def individual_employee_report(db, employees):
    """Generate individual employee attendance report"""
    st.subheader("👤 Individual Employee Report")
    
    # Employee selection
    employee_names = [f"{emp['name']} ({emp['employee_id']})" for emp in employees]
    selected_employee = st.selectbox("Select Employee", employee_names)
    
    # Get selected employee data
    selected_emp_data = None
    for emp in employees:
        if f"{emp['name']} ({emp['employee_id']})" == selected_employee:
            selected_emp_data = emp
            break
    
    if not selected_emp_data:
        st.error("Employee not found!")
        return
    
    # Date range selection
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input(
            "Start Date",
            value=date.today().replace(day=1),  # First day of current month
            max_value=date.today()
        )
    with col2:
        end_date = st.date_input(
            "End Date",
            value=date.today(),
            max_value=date.today()
        )
    
    if start_date > end_date:
        st.error("Start date cannot be after end date!")
        return
    
    # Generate report
    if st.button("📊 Generate Report", type="primary"):
        with st.spinner("Generating attendance report..."):
            # Get attendance data
            attendance_data = db.get_employee_attendance_data(
                selected_emp_data['id'], start_date, end_date
            )
            
            if not attendance_data:
                st.warning("No attendance data found for the selected period.")
                return
            
            # Analyze attendance
            analysis_results = AttendanceAnalyzer.analyze_employee_attendance(
                attendance_data, start_date, end_date
            )
            
            if not analysis_results:
                st.warning("No workdays found in the selected period.")
                return
            
            # Generate summary
            summary = AttendanceAnalyzer.generate_monthly_summary(analysis_results)
            
            # Display summary metrics
            st.markdown("### 📊 Summary Statistics")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Total Workdays", summary['total_workdays'])
                st.metric("Present Days", summary['present_days'])
            with col2:
                st.metric("Absent Days", summary['absent_days'])
                st.metric("Attendance Rate", f"{summary['attendance_rate']:.1f}%")
            with col3:
                st.metric("Avg Entry Delay", format_minutes_to_hms(summary['avg_entry_delay_minutes']))
                st.metric("Avg Early Exit", format_minutes_to_hms(summary['avg_exit_early_minutes']))
            
            # Display detailed daily report
            st.markdown("### 📅 Daily Attendance Details")
            
            # Create DataFrame for display
            daily_df = []
            for day in analysis_results:
                daily_df.append({
                    'Date': day['date'].strftime('%Y-%m-%d'),
                    'Day': day['date'].strftime('%A'),
                    'Status': day['status'],
                    'First Entry': day['first_entry'].strftime('%H:%M') if day['first_entry'] else 'N/A',
                    'Last Exit': day['last_exit'].strftime('%H:%M') if day['last_exit'] else 'N/A',
                    'Entry Delay': format_minutes_to_hms(day['entry_delay_minutes']) if day['entry_delay_minutes'] else '0 sn',
                    'Early Exit': format_minutes_to_hms(day['exit_early_minutes']) if day['exit_early_minutes'] else '0 sn'
                })
            
            df = pd.DataFrame(daily_df)
            
            # Color coding for status and delay/early exit
            def color_status(val):
                if val == 'Present':
                    return 'background-color: #d4edda'  # Light green
                elif val == 'Absent':
                    return 'background-color: #f8d7da'  # Light red
                elif 'Partial' in val:
                    return 'background-color: #fff3cd'  # Light yellow
                return ''
            def color_delay(val):
                try:
                    num = float(val.replace(',', '.').split()[0])
                    if num > 10:
                        return 'background-color: #f8d7da'  # Red for >10 min
                except:
                    pass
                return ''
            styled_df = df.style.applymap(color_status, subset=['Status']) \
                .applymap(color_delay, subset=['Entry Delay', 'Early Exit'])
            st.dataframe(styled_df, use_container_width=True)
            
            # Download options
            st.markdown("### 📥 Download Report")
            col1, col2 = st.columns(2)
            
            with col1:
                # CSV download
                csv = df.to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📄 Download as CSV",
                    data=csv,
                    file_name=f"attendance_report_{selected_emp_data['employee_id']}_{start_date}_{end_date}.csv",
                    mime="text/csv"
                )
            
            with col2:
                # Excel download with colors
                try:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Attendance Report"
                    
                    # Add title
                    ws['A1'] = f"Attendance Report - {selected_emp_data['name']} ({selected_emp_data['employee_id']})"
                    ws['A1'].font = Font(bold=True, size=14)
                    ws.merge_cells('A1:H1')
                    
                    # Add summary
                    ws['A3'] = "Summary Statistics"
                    ws['A3'].font = Font(bold=True, size=12)
                    
                    summary_data = [
                        ["Total Workdays", summary['total_workdays']],
                        ["Present Days", summary['present_days']],
                        ["Absent Days", summary['absent_days']],
                        ["Attendance Rate", f"{summary['attendance_rate']:.1f}%"],
                        ["Average Entry Delay", format_minutes_to_hms(summary['avg_entry_delay_minutes'])],
                        ["Average Early Exit", format_minutes_to_hms(summary['avg_exit_early_minutes'])]
                    ]
                    
                    for i, (label, value) in enumerate(summary_data):
                        ws[f'A{i+4}'] = label
                        ws[f'B{i+4}'] = value
                    
                    # Add daily data
                    ws['A13'] = "Daily Attendance Details"
                    ws['A13'].font = Font(bold=True, size=12)
                    
                    # Add headers
                    headers = ['Date', 'Day', 'Status', 'First Entry', 'Last Exit', 'Entry Delay', 'Early Exit']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=14, column=col, value=header)
                        cell.font = Font(bold=True)
                    
                    # Add data
                    for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 15):
                        for col_idx, value in enumerate(row_data, 1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                            # Apply colors based on status
                            if col_idx == 3:  # Status column
                                if value == 'Present':
                                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                                elif value == 'Absent':
                                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                                elif 'Partial' in value:
                                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                            # Entry Delay kırmızı (10 dakikadan fazlaysa)
                            if col_idx == 6:
                                try:
                                    # 'x saat y dakika z saniye' -> toplam saniye
                                    t = str(value)
                                    mins = 0
                                    h = re.search(r'(\d+) saat', t)
                                    m = re.search(r'(\d+) dakika', t)
                                    s = re.search(r'(\d+) saniye', t)
                                    if h:
                                        mins += int(h.group(1)) * 60
                                    if m:
                                        mins += int(m.group(1))
                                    if s:
                                        mins += int(s.group(1)) / 60
                                    if mins > 10:
                                        cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                                except:
                                    pass
                            # Early Exit kırmızı (10 dakikadan fazlaysa)
                            if col_idx == 7:
                                try:
                                    t = str(value)
                                    mins = 0
                                    h = re.search(r'(\d+) saat', t)
                                    m = re.search(r'(\d+) dakika', t)
                                    s = re.search(r'(\d+) saniye', t)
                                    if h:
                                        mins += int(h.group(1)) * 60
                                    if m:
                                        mins += int(m.group(1))
                                    if s:
                                        mins += int(s.group(1)) / 60
                                    if mins > 10:
                                        cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                                except:
                                    pass
                    
                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = None
                        for cell in column:
                            if not isinstance(cell, MergedCell):
                                if column_letter is None:
                                    column_letter = cell.column_letter
                                try:
                                    if cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                        if column_letter:
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # Save to BytesIO
                    excel_buffer = BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    
                    st.download_button(
                        label="📊 Download as Excel (with colors)",
                        data=excel_buffer.getvalue(),
                        file_name=f"attendance_report_{selected_emp_data['employee_id']}_{start_date}_{end_date}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                except ImportError:
                    st.warning("⚠️ Excel download requires openpyxl. Install with: `pip install openpyxl`")

def monthly_summary_report(db, employees):
    """Generate monthly summary report for all employees"""
    st.subheader("📊 Monthly Summary Report")
    
    # Month selection
    current_date = date.today()
    month_options = []
    for i in range(12):  # Last 12 months
        month_date = current_date - relativedelta(months=i)
        month_options.append(month_date.strftime('%B %Y'))
    
    selected_month = st.selectbox("Select Month", month_options)
    
    # Parse selected month
    month_date = datetime.strptime(selected_month, '%B %Y').date()
    start_date = month_date.replace(day=1)
    end_date = (start_date + relativedelta(months=1)) - timedelta(days=1)
    
    st.info(f"📅 Generating report for: {start_date.strftime('%B %Y')} ({start_date} to {end_date})")
    
    if st.button("📊 Generate Monthly Report", type="primary"):
        with st.spinner("Generating monthly summary report..."):
            # Get all employees attendance data
            all_attendance = db.get_all_employees_attendance(start_date, end_date)
            
            if not all_attendance:
                st.warning("No attendance data found for the selected month.")
                return
            
            # Group by employee
            employee_attendance = {}
            for record in all_attendance:
                emp_id = record['person_id']
                if emp_id not in employee_attendance:
                    employee_attendance[emp_id] = {
                        'name': record['name'],
                        'employee_id': record['employee_id'],
                        'department': record.get('department', 'N/A'),
                        'data': []
                    }
                employee_attendance[emp_id]['data'].append(record)
            
            # Analyze each employee
            summary_data = []
            for emp_id, emp_info in employee_attendance.items():
                analysis_results = AttendanceAnalyzer.analyze_employee_attendance(
                    emp_info['data'], start_date, end_date
                )
                
                if analysis_results:
                    summary = AttendanceAnalyzer.generate_monthly_summary(analysis_results)
                    summary_data.append({
                        'Employee Name': emp_info['name'],
                        'Employee ID': emp_info['employee_id'],
                        'Department': emp_info.get('department', 'N/A'),
                        'Total Workdays': summary['total_workdays'],
                        'Present Days': summary['present_days'],
                        'Absent Days': summary['absent_days'],
                        'Attendance Rate (%)': round(summary['attendance_rate'], 1),
                        'Avg Entry Delay': format_minutes_to_hms(summary['avg_entry_delay_minutes']),
                        'Avg Early Exit': format_minutes_to_hms(summary['avg_exit_early_minutes'])
                    })
            
            if not summary_data:
                st.warning("No valid attendance data found for the selected month.")
                return
            
            # Create DataFrame
            df = pd.DataFrame(summary_data)
            
            # Display summary
            st.markdown("### 📈 Monthly Summary")
            st.dataframe(df, use_container_width=True)
            
            # Download options
            st.markdown("### 📥 Download Monthly Report")
            col1, col2 = st.columns(2)
            
            with col1:
                csv = df.to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📄 Download as CSV",
                    data=csv,
                    file_name=f"monthly_summary_{start_date.strftime('%Y%m')}.csv",
                    mime="text/csv"
                )
            
            with col2:
                try:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Monthly Summary"
                    
                    # Add title
                    ws['A1'] = f"Monthly Attendance Summary - {start_date.strftime('%B %Y')}"
                    ws['A1'].font = Font(bold=True, size=14)
                    ws.merge_cells('A1:H1')
                    
                    # Add data
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    
                    # Style headers
                    for cell in ws[2]:  # Row 2 contains headers
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    
                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = None
                        for cell in column:
                            if not isinstance(cell, MergedCell):
                                if column_letter is None:
                                    column_letter = cell.column_letter
                                try:
                                    if cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                        if column_letter:
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                    
                    excel_buffer = BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    
                    st.download_button(
                        label="📊 Download as Excel",
                        data=excel_buffer.getvalue(),
                        file_name=f"monthly_summary_{start_date.strftime('%Y%m')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                except ImportError:
                    st.warning("⚠️ Excel download requires openpyxl. Install with: `pip install openpyxl`")

def department_report(db, employees):
    """Generate department-wise attendance report"""
    st.subheader("🏢 Department Report")
    
    # Get unique departments
    departments = list(set([emp['department'] for emp in employees if emp['department']]))
    departments.append('No Department')
    
    selected_department = st.selectbox("Select Department", departments)
    
    # Date range selection
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input(
            "Start Date",
            value=date.today().replace(day=1),
            max_value=date.today()
        )
    with col2:
        end_date = st.date_input(
            "End Date",
            value=date.today(),
            max_value=date.today()
        )
    
    if start_date > end_date:
        st.error("Start date cannot be after end date!")
        return
    
    if st.button("📊 Generate Department Report", type="primary"):
        with st.spinner("Generating department report..."):
            # Filter employees by department
            dept_employees = []
            for emp in employees:
                if selected_department == 'No Department':
                    if not emp['department']:
                        dept_employees.append(emp)
                else:
                    if emp['department'] == selected_department:
                        dept_employees.append(emp)
            
            if not dept_employees:
                st.warning(f"No employees found in department: {selected_department}")
                return
            
            # Get attendance data for all employees in department
            all_attendance = db.get_all_employees_attendance(start_date, end_date)
            
            # Filter by department employees
            dept_attendance = [record for record in all_attendance 
                             if any(emp['id'] == record['person_id'] for emp in dept_employees)]
            
            if not dept_attendance:
                st.warning("No attendance data found for the selected period.")
                return
            
            # Group by employee
            employee_attendance = {}
            for record in dept_attendance:
                emp_id = record['person_id']
                if emp_id not in employee_attendance:
                    employee_attendance[emp_id] = {
                        'name': record['name'],
                        'employee_id': record['employee_id'],
                        'department': record.get('department', 'N/A'),
                        'data': []
                    }
                employee_attendance[emp_id]['data'].append(record)
            
            # Analyze each employee
            summary_data = []
            for emp_id, emp_info in employee_attendance.items():
                analysis_results = AttendanceAnalyzer.analyze_employee_attendance(
                    emp_info['data'], start_date, end_date
                )
                
                if analysis_results:
                    summary = AttendanceAnalyzer.generate_monthly_summary(analysis_results)
                    summary_data.append({
                        'Employee Name': emp_info['name'],
                        'Employee ID': emp_info['employee_id'],
                        'Department': emp_info.get('department', 'N/A'),
                        'Total Workdays': summary['total_workdays'],
                        'Present Days': summary['present_days'],
                        'Absent Days': summary['absent_days'],
                        'Attendance Rate (%)': round(summary['attendance_rate'], 1),
                        'Avg Entry Delay': format_minutes_to_hms(summary['avg_entry_delay_minutes']),
                        'Avg Early Exit': format_minutes_to_hms(summary['avg_exit_early_minutes'])
                    })
            
            if not summary_data:
                st.warning("No valid attendance data found for the selected period.")
                return
            
            # Create DataFrame
            df = pd.DataFrame(summary_data)
            
            # Display summary
            st.markdown(f"### 📊 {selected_department} Department Summary")
            st.dataframe(df, use_container_width=True)
            
            # Department statistics
            st.markdown("### 📈 Department Statistics")
            col1, col2, col3 = st.columns(3)
            
            avg_attendance = df['Attendance Rate (%)'].mean()
            with col1:
                st.metric("Avg Attendance Rate", f"{avg_attendance:.1f}%")
            with col2:
                st.metric("Avg Entry Delay", df['Avg Entry Delay'].iloc[0] if not df.empty else "-")
            with col3:
                st.metric("Avg Early Exit", df['Avg Early Exit'].iloc[0] if not df.empty else "-")
            
            # Download options
            st.markdown("### 📥 Download Department Report")
            col1, col2 = st.columns(2)
            
            with col1:
                csv = df.to_csv(index=False, encoding='utf-8-sig')
                st.download_button(
                    label="📄 Download as CSV",
                    data=csv,
                    file_name=f"department_report_{selected_department}_{start_date}_{end_date}.csv",
                    mime="text/csv"
                )
            
            with col2:
                try:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Department Report"
                    
                    # Add title
                    ws['A1'] = f"Department Report - {selected_department} ({start_date} to {end_date})"
                    ws['A1'].font = Font(bold=True, size=14)
                    ws.merge_cells('A1:H1')
                    
                    # Add data
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    
                    # Style headers
                    for cell in ws[2]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    
                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = None
                        for cell in column:
                            if not isinstance(cell, MergedCell):
                                if column_letter is None:
                                    column_letter = cell.column_letter
                                try:
                                    if cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                        if column_letter:
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                    
                    excel_buffer = BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    
                    st.download_button(
                        label="📊 Download as Excel",
                        data=excel_buffer.getvalue(),
                        file_name=f"department_report_{selected_department}_{start_date}_{end_date}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                except ImportError:
                    st.warning("⚠️ Excel download requires openpyxl. Install with: `pip install openpyxl`")

def face_recognition_test_page(db):
    """Page for testing face recognition with camera only"""
    st.header("🔍 Face Recognition Test (Kamera)")

    st.info("Kameradan yüz tanıma için kameranızı açın. Yüzünüz algılandığında otomatik olarak tanıma yapılacaktır.")

    # Camera location selection
    st.markdown("### 📍 Kamera Konumu Seçimi")
    camera_options = {info["name"]: key for key, info in CAMERA_LOCATIONS.items()}
    selected_camera = st.selectbox(
        "Hangi kamera konumunu test ediyorsunuz?",
        options=list(camera_options.keys()),
        index=0,
        help="Bu seçim, erişim loglarında hangi kapı/kamera olarak kaydedileceğini belirler"
    )
    
    selected_camera_key = camera_options[selected_camera]
    camera_info = CAMERA_LOCATIONS[selected_camera_key]
    
    # Show camera info
    col1, col2, col3 = st.columns(3)
    with col1:
        st.info(f"**Konum:** {camera_info['name']}")
    with col2:
        st.info(f"**Erişim Tipi:** {camera_info['access_type']}")
    with col3:
        st.info(f"**Açıklama:** {camera_info['description']}")

    # VideoTransformer'a db nesnesini ve kamera konumunu ilet
    def video_transformer_factory():
        vt = VideoTransformer(camera_location=selected_camera_key)
        vt.set_db(db)
        return vt

    webrtc_ctx = webrtc_streamer(
        key=f"face-recognition-test-{selected_camera_key}",
        video_processor_factory=video_transformer_factory,
        rtc_configuration=RTCConfiguration(
            {"iceServers": [{"urls": ["stun:stun.l.google.com:19302"]}]}
        ),
        media_stream_constraints={
            "video": {"width": 640, "height": 480},
            "audio": False
        }
    )

    if webrtc_ctx.video_processor:
        vt = webrtc_ctx.video_processor
        if vt.last_recognition:
            person = vt.last_recognition
            confidence = vt.last_confidence
            access_type = camera_info["access_type"]
            
            # Show appropriate message based on access type
            if access_type == "entry":
                st.success(f"✅ Giris Izni Verildi: {person['name']} ({confidence:.1%} guven)")
            elif access_type == "exit":
                st.warning(f"🚪 Cikis Izni Verildi: {person['name']} ({confidence:.1%} guven)")
            else:  # internal
                st.info(f"🔓 Ic Kapi Acildi: {person['name']} ({confidence:.1%} güven)")
            
            col1, col2 = st.columns(2)
            with col1:
                if person['face_image']:
                    try:
                        stored_image = FaceProcessor.base64_to_image(person['face_image'])
                        if stored_image:
                            st.image(stored_image, width=200, caption="Kayıtlı Fotoğraf")
                    except:
                        st.write("Fotoğraf yok")
            with col2:
                st.write(f"**Ad:** {person['name']}")
                st.write(f"**Çalışan No:** {person['employee_id']}")
                st.write(f"**Departman:** {person['department'] or 'N/A'}")
                st.write(f"**Güven:** {confidence:.1%}")
                st.write(f"**Konum:** {camera_info['name']}")
                st.write(f"**Erişim Tipi:** {access_type}")
        else:
            st.info("Kamerada yüz algılandığında ve tanındığında burada bilgi gözükecek.")

if __name__ == "__main__":
    main()